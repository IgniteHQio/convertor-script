import streamlit as st
import json
import re
import pandas as pd
import requests
from bs4 import BeautifulSoup
from deep_translator import GoogleTranslator
from openpyxl import load_workbook
from openpyxl.styles import PatternFill
import io

# --- Configuration ---
APP_PASSWORD = "Abcd@1234"

def check_password():
    if "password_correct" not in st.session_state:
        st.session_state["password_correct"] = False
    if st.session_state["password_correct"]: return True
    st.title("🔒 Access Restricted")
    password_input = st.text_input("Enter App Password", type="password")
    if st.button("Unlock"):
        if password_input == APP_PASSWORD:
            st.session_state["password_correct"] = True
            st.rerun()
        else:
            st.error("🚫 Incorrect password")
    return False

def split_text(text):
    if not text: return "", ""
    text = str(text).strip()
    en, ar = "", ""
    if '|' in text:
        parts = [p.strip() for p in text.split('|')]
        for p in parts:
            if re.search(r'[\u0600-\u06FF]', p): ar = p
            else: en = p
    else:
        en_match = re.findall(r'[a-zA-Z0-9\s&\'\.]+', text)
        ar_match = re.findall(r'[\u0600-\u06FF\s]+', text)
        en = " ".join([m.strip() for m in en_match if m.strip()]).strip()
        ar = " ".join([m.strip() for m in ar_match if m.strip()]).strip()
    return en, ar

def process_translation(en_val, ar_val):
    t_en, t_ar = False, False
    if ar_val and not en_val:
        try:
            en_val = GoogleTranslator(source='ar', target='en').translate(ar_val)
            t_en = True
        except: pass
    elif en_val and not ar_val:
        try:
            ar_val = GoogleTranslator(source='en', target='ar').translate(en_val)
            t_ar = True
        except: pass
    return en_val, ar_val, t_en, t_ar

def find_key_recursive(data, key_name):
    if isinstance(data, dict):
        if key_name in data: return data[key_name]
        for v in data.values():
            res = find_key_recursive(v, key_name)
            if res: return res
    elif isinstance(data, list):
        for i in data:
            res = find_key_recursive(i, key_name)
            if res: return res
    return None

def fetch_full_salon_data(salon_url):
    headers = {'User-Agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/120.0.0.0 Safari/537.36'}
    try:
        base_res = requests.get(salon_url, headers=headers, timeout=10)
        soup = BeautifulSoup(base_res.text, 'html.parser')
        next_data_script = soup.find('script', id='__NEXT_DATA__')
        if not next_data_script: return None, "Could not find page data script."
        
        full_page_json = json.loads(next_data_script.string)
        build_id = full_page_json.get('buildId')
        
        match = re.search(r'/a/([^/?#]+)', salon_url)
        if not match: return None, "Invalid URL handle."
        handle = match.group(1)
        
        json_url = f"https://www.fresha.com/_next/data/{build_id}/a/{handle}.json"
        json_res = requests.get(json_url, headers=headers, timeout=10)
        service_json = json_res.json()
        
        return {"page_json": full_page_json, "service_json": service_json}, None
    except Exception as e:
        return None, str(e)

if check_password():
    st.set_page_config(page_title="SALON JSON to EXCEL", page_icon="✂️")
    st.title("✂️ SALON DATA SCRAPER")

    if "master_data" not in st.session_state:
        st.session_state["master_data"] = None

    url_input = st.text_input("Paste Fresha Salon Homepage URL:")
    if st.button("Fetch Salon & Team Data"):
        data, err = fetch_full_salon_data(url_input)
        if err: st.error(err)
        else:
            st.session_state["master_data"] = data
            st.success("✅ Successfully fetched Service and Team data!")

    if st.session_state["master_data"]:
        master = st.session_state["master_data"]
        
        loc_info = find_key_recursive(master['service_json'], 'location') or {}
        menu_data = find_key_recursive(master['service_json'], 'services') or find_key_recursive(master['service_json'], 'categories')
        employee_data = find_key_recursive(master['page_json'], 'employeeProfiles')
        
        team_rows_raw = []
        if employee_data and 'edges' in employee_data:
            for edge in employee_data['edges']:
                node = edge.get('node', {})
                team_rows_raw.append(node)

        st.info(f"Salon: **{loc_info.get('name')}** | Menu: {len(menu_data) if menu_data else 0} Groups | Team: {len(team_rows_raw)}")

        if st.button("🚀 Generate Final Excel"):
            # --- Process Team Data with Translation ---
            processed_team = []
            if team_rows_raw:
                t_prog = st.progress(0)
                for idx, node in enumerate(team_rows_raw):
                    t_prog.progress((idx + 1) / len(team_rows_raw))
                    
                    # Process Name
                    n_en, n_ar, _, _ = process_translation(*split_text(node.get('displayName')))
                    # Process Job Title
                    j_en, j_ar, _, _ = process_translation(*split_text(node.get('jobTitle')))
                    
                    processed_team.append({
                        "Staff ID": node.get('employeeId'),
                        "Name (EN)": n_en,
                        "Name (AR)": n_ar,
                        "Job Title (EN)": j_en,
                        "Job Title (AR)": j_ar,
                        "Avatar URL": node.get('avatar', {}).get('url') if node.get('avatar') else "No Image"
                    })

            # --- Process Service Items ---
            items_list, cell_highlights = [], []
            
            # --- UPDATED: Added Address Field ---
            info_rows = [
                {"Field": "Name", "Value": loc_info.get('name')},
                {"Field": "Address", "Value": loc_info.get('address')},
                {"Field": "Description", "Value": loc_info.get('description')},
                {"Field": "Contact Number", "Value": loc_info.get('contactNumber')},
                {"Field": "Cover Image", "Value": loc_info.get('coverImage', {}).get('url')}
            ]

            if menu_data:
                all_items = [(g.get('name', ''), i) for g in menu_data for i in g.get('items', [])]
                prog = st.progress(0)
                for idx, (g_name, item) in enumerate(all_items):
                    prog.progress((idx + 1) / len(all_items))
                    c_en, c_ar, ce_t, ca_t = process_translation(*split_text(g_name))
                    i_en, i_ar, ie_t, ia_t = process_translation(*split_text(item.get('name', '')))
                    d_en, d_ar, de_t, da_t = process_translation(*split_text(item.get('description') or ""))
                    
                    price = item.get('formattedRetailPrice') or item.get('price', {}).get('formatted', '')
                    duration = item.get('caption', '')
                    service_id = item.get('id', '')

                    row_num = len(items_list) + 2
                    h = []
                    if ce_t: h.append(2)
                    if ca_t: h.append(3)
                    if ie_t: h.append(4)
                    if ia_t: h.append(5)
                    if de_t: h.append(6)
                    if da_t: h.append(7)
                    if h: cell_highlights.append((row_num, h))

                    items_list.append({
                        "Service ID": service_id,
                        "Category (EN)": c_en, 
                        "Category (AR)": c_ar,
                        "Service (EN)": i_en, 
                        "Service (AR)": i_ar,
                        "Desc (EN)": d_en, 
                        "Desc (AR)": d_ar,
                        "Price": price, 
                        "DURATION": duration
                    })

            output = io.BytesIO()
            with pd.ExcelWriter(output, engine='openpyxl') as writer:
                pd.DataFrame(info_rows).to_excel(writer, sheet_name='INFO', index=False)
                pd.DataFrame(items_list).to_excel(writer, sheet_name='ITEMS', index=False)
                pd.DataFrame(processed_team).to_excel(writer, sheet_name='TEAM', index=False)

            output.seek(0)
            wb = load_workbook(output)
            ws = wb['ITEMS']
            yellow = PatternFill(start_color='FFFF00', end_color='FFFF00', fill_type='solid')
            for r_idx, cols in cell_highlights:
                for c_idx in cols: ws.cell(row=r_idx, column=c_idx).fill = yellow
            
            final_out = io.BytesIO()
            wb.save(final_out)
            st.download_button("📥 Download Final Excel", final_out.getvalue(), f"{loc_info.get('name','salon')}.xlsx")

    if st.button("🧹 Reset"):
        st.session_state["master_data"] = None
        st.rerun()